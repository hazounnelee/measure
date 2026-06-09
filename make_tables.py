#!/usr/bin/env python3
"""batch_summary.json(활물질/대입경/소입경)으로부터 기초통계량 및 등급화 Excel 표 생성."""
import argparse
import json
import math
import shutil
import sys
from pathlib import Path

# 지표별 (폴더명, per-image 값 키, batch 값 키, reverse)
_GRADE_METRICS = [
    ("입도_표준편차", "particle_size_std_um",         "particle_size_um.std",           False),
    ("타원도",        "particle_sphericity_prime_mean","particle_sphericity_prime.mean",  True),
    ("구형도",        "particle_sphericity_mean",      "particle_sphericity.mean",        True),
    ("미분_깨짐",     "fine_particle_ratio_percent",   "fine_particle_ratio_percent",     False),
]

_ROI_CANDIDATES        = ["input_roi.png", "02_input_roi.png"]
_CLASSIFIED_CANDIDATES = ["classified.png", "06_pipeline_classified.png"]

_TEMPLATE = Path(__file__).parent / "tables.xlsx"

# 입도 RMSD 기준 (µm)
_REF_ACTIVE = None
_REF_LARGE = 10.0
_REF_SMALL = 4.0

# 이미지 하단 바 설정
_FONT_KO   = "/System/Library/Fonts/AppleSDGothicNeo.ttc"
_FONT_SIZE = 14
_BAR_PAD   = 7
_LINE_GAP  = 3


def _load(path: str | None) -> dict | None:
    if path is None:
        return None
    p = Path(path)
    if not p.exists():
        print(f"[ERROR] 파일 없음: {p}", file=sys.stderr)
        sys.exit(1)
    with p.open(encoding="utf-8") as f:
        return json.load(f)


def _get(d: dict, key: str) -> float | None:
    """점(.) 구분 중첩 키 또는 단순 키로 float 값 반환. 없거나 None이면 None 반환."""
    v = d
    for part in key.split("."):
        if not isinstance(v, dict):
            return None
        v = v.get(part)
        if v is None:
            return None
    return float(v)


def _safe_float(v) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _quartiles(
    d1: dict | None, d2: dict | None, d3: dict | None, per_image_key: str
) -> tuple[float, float, float]:
    """세 배치의 per-image 값을 모아 Q1/Q2/Q3 반환."""
    import numpy as np
    vals = []
    for d in (d1, d2, d3):
        if d is None:
            continue
        for img_id_entry in d.get("img_ids", []):
            for fe in img_id_entry.get("files", []):
                v = fe.get(per_image_key)
                if v is not None:
                    vals.append(float(v))
    if not vals:
        return 0.0, 0.0, 0.0
    return (
        float(np.percentile(vals, 25)),
        float(np.percentile(vals, 50)),
        float(np.percentile(vals, 75)),
    )


def _lot_stats(d: dict | None, key: str) -> tuple[float | None, float | None, float | None]:
    """per-file 값을 모아 (mean, median, std) 반환."""
    import numpy as np
    vals = []
    for dict_g in (d or {}).get("img_ids", []):
        for dict_f in dict_g.get("files", []):
            v = _safe_float(dict_f.get(key))
            if v is not None:
                vals.append(v)
    if not vals:
        return None, None, None
    a = np.array(vals)
    return (
        float(np.mean(a)),
        float(np.median(a)),
        float(np.std(a, ddof=1)) if len(vals) > 1 else 0.0,
    )


def _lot_rmsd_stats(
    d: dict | None,
    mean_key: str,
    std_key: str,
    ref: float,
) -> tuple[float | None, float | None, float | None]:
    """per-file RMSD = sqrt(std²+(mean-ref)²) 를 모아 (mean, median, std) 반환."""
    import numpy as np
    vals = []
    for dict_g in (d or {}).get("img_ids", []):
        for dict_f in dict_g.get("files", []):
            fv_mean = _safe_float(dict_f.get(mean_key))
            if fv_mean is None:
                continue
            fv_std = _safe_float(dict_f.get(std_key))
            var = fv_std ** 2 if fv_std is not None else 0.0
            vals.append(math.sqrt(var + (fv_mean - ref) ** 2))
    if not vals:
        return None, None, None
    a = np.array(vals)
    return (
        float(np.mean(a)),
        float(np.median(a)),
        float(np.std(a, ddof=1)) if len(vals) > 1 else 0.0,
    )


def _write_row(ws, row: int, mean_: float | None, med_: float | None, std_: float | None) -> None:
    if mean_ is None:
        return
    ws.cell(row=row, column=3).value = round(mean_, 3)
    ws.cell(row=row, column=4).value = round(med_,  3)
    ws.cell(row=row, column=5).value = round(std_,  3)


def _grade_quartile(
    value: float, q1: float, q2: float, q3: float, reverse: bool = False
) -> int:
    """4분위수 기반 4등급 반환 (1=최우수)."""
    if not reverse:
        if value <= q1: return 1
        if value <= q2: return 2
        if value <= q3: return 3
        return 4
    else:
        if value >= q3: return 1
        if value >= q2: return 2
        if value >= q1: return 3
        return 4


def _grade_min(
    g: int, q1: float, q2: float, q3: float, reverse: bool
) -> float | None:
    """등급 g의 최솟값(하한). 하한이 없으면 None."""
    if not reverse:
        return [None, q1, q2, q3][g - 1]
    else:
        return [q3, q2, q1, None][g - 1]


def make_tables(
    d_active: dict | None,
    d_large: dict | None,
    d_small: dict | None,
    path_template: Path,
    path_output: Path,
) -> None:
    import openpyxl

    shutil.copy2(path_template, path_output)
    wb = openpyxl.load_workbook(path_output)

    # ── 기초통계량 (개별 입자) ────────────────────────────────────────────
    ws_ind = wb["기초통계량 및 처리 시간 (개별 입자)"]
    stat_rows_ind = [
        (3,  4,  5,  "particle_size_um"),
        (6,  7,  8,  "particle_sphericity_prime"),
        (9,  10, 11, "particle_sphericity"),
        (12, 13, 14, "fine_particle_ratio_percent_stats"),
    ]
    for row_a, row_l, row_s, key in stat_rows_ind:
        for row, d in ((row_a, d_active), (row_l, d_large), (row_s, d_small)):
            if d is None:
                continue
            s = d.get(key)
            if s is None:
                continue
            _write_row(ws_ind, row, float(s["mean"]), float(s["median"]), float(s["std"]))

    # ── 기초통계량 (LOT 평균) ─────────────────────────────────────────────
    ws_lot = wb["기초통계량 및 처리 시간 (LOT)"]

    # 입도 (µm)
    for row, d in ((3, d_active), (4, d_large), (5, d_small)):
        _write_row(ws_lot, row, *_lot_stats(d, "particle_mean_size_um"))

    # 입도 표준편차 (µm)
    for row, d in ((6, d_active), (7, d_large), (8, d_small)):
        _write_row(ws_lot, row, *_lot_stats(d, "particle_size_std_um"))

    # 입도 RMSD (µm)
    if _REF_ACTIVE is not None:
        _write_row(ws_lot, 9, *_lot_rmsd_stats(d_active, "particle_mean_size_um", "particle_size_std_um", _REF_ACTIVE))
    _write_row(ws_lot, 10, *_lot_rmsd_stats(d_large, "particle_mean_size_um", "particle_size_std_um", _REF_LARGE))
    _write_row(ws_lot, 11, *_lot_rmsd_stats(d_small, "particle_mean_size_um", "particle_size_std_um", _REF_SMALL))

    # 타원도
    for row, d in ((12, d_active), (13, d_large), (14, d_small)):
        _write_row(ws_lot, row, *_lot_stats(d, "particle_sphericity_prime_mean"))

    # 구형도
    for row, d in ((15, d_active), (16, d_large), (17, d_small)):
        _write_row(ws_lot, row, *_lot_stats(d, "particle_sphericity_mean"))

    # 미분/깨짐 비율
    for row, d in ((18, d_active), (19, d_large), (20, d_small)):
        _write_row(ws_lot, row, *_lot_stats(d, "fine_particle_ratio_percent"))

    # 이미지 당 처리 시간
    for row, d in ((21, d_active), (22, d_large), (23, d_small)):
        if d is None:
            continue
        s = d.get("processing_time_sec")
        if s is None:
            continue
        _write_row(ws_lot, row, float(s["mean"]), float(s["median"]), float(s["std"]))

    # ── 등급화 ────────────────────────────────────────────────────────────
    ws_grade = wb["등급화"]

    q1_size, q2_size, q3_size = _quartiles(d_active, d_large, d_small, "particle_size_std_um")
    q1_ell,  q2_ell,  q3_ell  = _quartiles(d_active, d_large, d_small, "particle_sphericity_prime_mean")
    q1_sph,  q2_sph,  q3_sph  = _quartiles(d_active, d_large, d_small, "particle_sphericity_mean")
    q1_frag, q2_frag, q3_frag = _quartiles(d_active, d_large, d_small, "fine_particle_ratio_percent")

    grade_rows = [
        (4,  5,  6,  "particle_size_um.std",            q1_size, q2_size, q3_size, False),
        (8,  9,  10, "particle_sphericity_prime.mean",   q1_ell,  q2_ell,  q3_ell,  True),
        (12, 13, 14, "particle_sphericity.mean",          q1_sph,  q2_sph,  q3_sph,  True),
        (16, 17, 18, "fine_particle_ratio_percent",       q1_frag, q2_frag, q3_frag, False),
    ]
    for row_a, row_l, row_s, val_key, q1, q2, q3, reverse in grade_rows:
        for row, d in ((row_a, d_active), (row_l, d_large), (row_s, d_small)):
            if d is None:
                continue
            value = _get(d, val_key)
            if value is None:
                continue
            g = _grade_quartile(value, q1, q2, q3, reverse=reverse)
            min_val = _grade_min(g, q1, q2, q3, reverse)
            if min_val is not None:
                ws_grade.cell(row=row, column=g + 2).value = round(min_val, 4)

    wb.save(path_output)
    print(f"[done] {path_output}")


def _find(output_dir: Path, candidates: list[str]) -> Path | None:
    for name in candidates:
        p = output_dir / name
        if p.exists():
            return p
    return None


def _annotate_image(path_src: Path, path_dst: Path, lines: list[str]) -> None:
    """이미지 아래 검정 바를 추가하고 지표 텍스트를 씀."""
    from PIL import Image, ImageDraw, ImageFont
    img = Image.open(path_src).convert("RGB")
    w, h = img.size

    line_h = _FONT_SIZE + _LINE_GAP
    bar_h  = _BAR_PAD * 2 + line_h * len(lines)

    canvas = Image.new("RGB", (w, h + bar_h), (0, 0, 0))
    canvas.paste(img, (0, 0))
    draw = ImageDraw.Draw(canvas)

    try:
        font = ImageFont.truetype(_FONT_KO, _FONT_SIZE)
    except (IOError, OSError):
        font = ImageFont.load_default()

    y = h + _BAR_PAD
    for line in lines:
        draw.text((_BAR_PAD, y), line, fill=(210, 210, 210), font=font)
        y += line_h

    canvas.save(path_dst)


def _build_metric_lines(
    fe: dict,
    refs: dict,
    rmsd_ref: float | None,
) -> list[str]:
    """per-file 지표를 텍스트 라인 리스트로 반환."""
    lines = []

    mean_um = _safe_float(fe.get("particle_mean_size_um"))
    std_um  = _safe_float(fe.get("particle_size_std_um"))

    if mean_um is not None:
        s = f"입도: {mean_um:.2f}"
        s += f" +/- {std_um:.2f} um" if std_um is not None else " um"
        lines.append(s)

    if std_um is not None:
        q1, q2, q3, reverse = refs["입도_표준편차"]
        g = _grade_quartile(std_um, q1, q2, q3, reverse)
        lines.append(f"입도 표준편차: {std_um:.2f} um  [등급 {g}]")

    if mean_um is not None and rmsd_ref is not None:
        var  = std_um ** 2 if std_um is not None else 0.0
        rmsd = math.sqrt(var + (mean_um - rmsd_ref) ** 2)
        lines.append(f"입도 RMSD: {rmsd:.2f} um  (기준 {rmsd_ref:.0f} um)")

    ell = _safe_float(fe.get("particle_sphericity_prime_mean"))
    if ell is not None:
        q1, q2, q3, reverse = refs["타원도"]
        g = _grade_quartile(ell, q1, q2, q3, reverse)
        lines.append(f"타원도: {ell:.4f}  [등급 {g}]")

    sph = _safe_float(fe.get("particle_sphericity_mean"))
    if sph is not None:
        q1, q2, q3, reverse = refs["구형도"]
        g = _grade_quartile(sph, q1, q2, q3, reverse)
        lines.append(f"구형도: {sph:.4f}  [등급 {g}]")

    frag = _safe_float(fe.get("fine_particle_ratio_percent"))
    if frag is not None:
        q1, q2, q3, reverse = refs["미분_깨짐"]
        g = _grade_quartile(frag, q1, q2, q3, reverse)
        lines.append(f"미분/깨짐: {frag:.2f}%  [등급 {g}]")

    return lines


def export_grade_images(
    d: dict,
    label: str,
    path_outdir: Path,
    q1_size: float, q2_size: float, q3_size: float,
    q1_ell: float,  q2_ell: float,  q3_ell: float,
    q1_sph: float,  q2_sph: float,  q3_sph: float,
    q1_frag: float, q2_frag: float, q3_frag: float,
) -> None:
    """배치 내 각 이미지를 지표별 등급 폴더에 저장 (하단 바에 전 지표 표시)."""
    refs = {
        "입도_표준편차": (q1_size, q2_size, q3_size, False),
        "타원도":        (q1_ell,  q2_ell,  q3_ell,  True),
        "구형도":        (q1_sph,  q2_sph,  q3_sph,  True),
        "미분_깨짐":     (q1_frag, q2_frag, q3_frag, False),
    }
    rmsd_ref = {"active": _REF_ACTIVE, "large": _REF_LARGE, "small": _REF_SMALL}.get(label)

    copied = 0
    for img_id_entry in d.get("img_ids", []):
        for file_entry in img_id_entry.get("files", []):
            out_src  = Path(file_entry.get("output_dir", ""))
            img_name = Path(file_entry.get("image_name", file_entry.get("input_path", "unknown"))).stem

            path_roi        = _find(out_src, _ROI_CANDIDATES)
            path_classified = _find(out_src, _CLASSIFIED_CANDIDATES)
            if path_roi is None or path_classified is None:
                print(f"  [skip] 이미지 없음: {out_src}")
                continue

            lines = _build_metric_lines(file_entry, refs, rmsd_ref)

            for metric_name, val_key, _, _ in _GRADE_METRICS:
                val = file_entry.get(val_key)
                if val is None:
                    continue
                q1, q2, q3, reverse = refs[metric_name]
                g = _grade_quartile(float(val), q1, q2, q3, reverse=reverse)

                dest = path_outdir / label / metric_name / f"grade_{g}" / img_name
                dest.mkdir(parents=True, exist_ok=True)
                _annotate_image(path_roi,        dest / "input_roi.png",  lines)
                _annotate_image(path_classified, dest / "classified.png", lines)
                copied += 1

    print(f"  [{label}] {copied}개 이미지 저장 완료 → {path_outdir / label}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="batch_summary.json → 기초통계량/등급화 Excel 표 생성",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--active", metavar="JSON", default=None,
                        help="활물질 batch_summary.json 경로")
    parser.add_argument("--large", metavar="JSON", default=None,
                        help="대입경 batch_summary.json 경로")
    parser.add_argument("--small", metavar="JSON", default=None,
                        help="소입경 batch_summary.json 경로")
    parser.add_argument("--template", metavar="XLSX", default=str(_TEMPLATE),
                        help="표 템플릿 xlsx 경로")
    parser.add_argument("-o", "--output", metavar="XLSX", default="batch_tables.xlsx",
                        help="출력 xlsx 경로")
    parser.add_argument("--grade-images", metavar="DIR", default=None,
                        help="등급별 이미지 쌍(input_roi+classified)을 내보낼 최상위 디렉토리")
    args = parser.parse_args()

    if args.active is None and args.large is None and args.small is None:
        parser.error("--active / --large / --small 중 하나 이상 지정하세요.")

    d_active = _load(args.active)
    d_large = _load(args.large)
    d_small = _load(args.small)

    path_template = Path(args.template)
    if not path_template.exists():
        print(f"[ERROR] 템플릿 없음: {path_template}", file=sys.stderr)
        sys.exit(1)

    make_tables(d_active, d_large, d_small, path_template, Path(args.output))

    if args.grade_images:
        path_gi = Path(args.grade_images)
        q1_size, q2_size, q3_size = _quartiles(d_active, d_large, d_small, "particle_size_std_um")
        q1_ell,  q2_ell,  q3_ell  = _quartiles(d_active, d_large, d_small, "particle_sphericity_prime_mean")
        q1_sph,  q2_sph,  q3_sph  = _quartiles(d_active, d_large, d_small, "particle_sphericity_mean")
        q1_frag, q2_frag, q3_frag = _quartiles(d_active, d_large, d_small, "fine_particle_ratio_percent")
        print("[grade-images]")
        if d_active is not None:
            export_grade_images(d_active, "active", path_gi,
                                q1_size, q2_size, q3_size,
                                q1_ell,  q2_ell,  q3_ell,
                                q1_sph,  q2_sph,  q3_sph,
                                q1_frag, q2_frag, q3_frag)
        if d_large is not None:
            export_grade_images(d_large, "large", path_gi,
                                q1_size, q2_size, q3_size,
                                q1_ell,  q2_ell,  q3_ell,
                                q1_sph,  q2_sph,  q3_sph,
                                q1_frag, q2_frag, q3_frag)
        if d_small is not None:
            export_grade_images(d_small, "small", path_gi,
                                q1_size, q2_size, q3_size,
                                q1_ell,  q2_ell,  q3_ell,
                                q1_sph,  q2_sph,  q3_sph,
                                q1_frag, q2_frag, q3_frag)


if __name__ == "__main__":
    main()
